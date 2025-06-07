from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import math
import os

# 参数设置
img_width, img_height = 255, 127
lines_per_image = 8  # 每张图显示8行（4个单词，每个单词2行）
font_size = 12

# 字体路径（注意转义或使用原始字符串）
font_en_path = r"H:\下载\pic\consola.ttf"
font_zh_path = r"H:\下载\pic\原版宋体.ttf"

# 加载字体
try:
    font_en = ImageFont.truetype(font_en_path, font_size)
    font_zh = ImageFont.truetype(font_zh_path, font_size)
except Exception as e:
    print(f"字体加载失败：{e}")
    font_en = ImageFont.load_default()
    font_zh = ImageFont.load_default()

# Excel 数据读取
excel_path = "words.xlsx"  # 替换为你的 Excel 文件名
wb = load_workbook(excel_path)
ws = wb.active

# 提取 A 列和 B 列（单词和释义）
words = []
for row in ws.iter_rows(min_row=1, values_only=True):
    if row[0] and row[1]:
        words.append((str(row[0]), str(row[1])))

# 创建输出文件夹
output_dir = "output_images"
os.makedirs(output_dir, exist_ok=True)

# 每张图显示4组数据
group_size = 4
num_images = math.ceil(len(words) / group_size)

# 生成图像
for i in range(num_images):
    chunk = words[i*group_size:(i+1)*group_size]

    # 创建图像
    image = Image.new("RGB", (img_width, img_height), "white")
    draw = ImageDraw.Draw(image)

    # 每行高度
    line_height = img_height // lines_per_image

    # 写入文本
    for j, (word, meaning) in enumerate(chunk):
        y_word = j * 2 * line_height
        y_meaning = y_word + line_height

        draw.text((5, y_word), word, fill="black", font=font_en)
        draw.text((5, y_meaning), meaning, fill="gray", font=font_zh)

    # 保存为 jpg，文件名为 1.jpg, 2.jpg ...
    filename = f"{i+1}.jpg"
    image.save(os.path.join(output_dir, filename), format="JPEG")

print(f"共生成 {num_images} 张图片，保存在 {output_dir}/ 目录下")
